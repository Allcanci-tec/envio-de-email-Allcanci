#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para exportar TODOS os pedidos de Correios com dados estruturados.
Formatos: JSON, CSV e Excel (se disponível).
"""

import sys
import json
import csv
from datetime import datetime
from bling_auth import bling_get, get_token_info

def main():
    print("\n" + "="*130)
    print("  EXPORTAR PEDIDOS CORREIOS")
    print("="*130 + "\n")
    
    CORREIOS_ID = 151483
    
    # 1. Verificar token
    print("1. VERIFICANDO TOKEN")
    print("-" * 130)
    info = get_token_info()
    print(f"  Status: {info['status']} | Expira em {info['expires_in_minutes']} minutos\n")
    
    # 2. Buscar pedidos
    print("2. BUSCANDO PEDIDOS CORREIOS")
    print("-" * 130)
    
    try:
        print("  GET /pedidos/vendas?logistica=151483&limite=1000\n")
        response = bling_get("/pedidos/vendas", params={"logistica": CORREIOS_ID, "limite": 1000})
        
        pedidos_lista = response.get("data", [])
        
        if not pedidos_lista:
            print("  Nenhum pedido encontrado!\n")
            return 1
        
        print(f"  Total de pedidos encontrados: {len(pedidos_lista)}\n")
        
        # 3. Processar cada pedido
        print("3. PROCESSANDO PEDIDOS")
        print("-" * 130)
        print("  Buscando detalhes de cada pedido...\n")
        
        pedidos_processados = []
        
        for idx, pedido_resumido in enumerate(pedidos_lista, 1):
            # Barra de progresso
            if idx % 10 == 0:
                print(f"    Processado: {idx}/{len(pedidos_lista)}")
            
            try:
                pedido_id = pedido_resumido.get("id")
                
                # Pegar detalhes completos
                pedido_completo = bling_get(f"/pedidos/vendas/{pedido_id}")
                pedido_data = pedido_completo.get("data", {})
                
                # Extrair dados
                dados = {
                    "numero_pedido": pedido_data.get("numero"),
                    "id_pedido": pedido_id,
                    "data_pedido": pedido_data.get("data"),
                    "data_saida": pedido_data.get("dataSaida"),
                    "cliente_nome": pedido_data.get("contato", {}).get("nome"),
                    "cliente_tipo_pessoa": pedido_data.get("contato", {}).get("tipoPessoa"),
                    "cliente_documento": pedido_data.get("contato", {}).get("numeroDocumento"),
                    "valor_total": pedido_data.get("total"),
                    "valor_produtos": pedido_data.get("totalProdutos"),
                    "situacao_id": pedido_data.get("situacao", {}).get("id"),
                    "desconto": pedido_data.get("desconto", {}).get("valor"),
                    "observacoes": pedido_data.get("observacoes"),
                    "data_prevista": pedido_data.get("dataPrevista"),
                    "quantidade_itens": len(pedido_data.get("itens", [])),
                }
                
                pedidos_processados.append(dados)
                
            except Exception as e:
                print(f"    ⚠️  Erro ao processar pedido {pedido_id}: {e}")
                continue
        
        print(f"\n  ✅ Total de pedidos processados: {len(pedidos_processados)}\n")
        
        # 4. Exportar JSON
        print("4. EXPORTANDO DADOS")
        print("-" * 130)
        
        json_file = "pedidos_correios_exportacao.json"
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump({
                "timestamp": datetime.now().isoformat(),
                "transportadora": "Correios",
                "logistica_id": CORREIOS_ID,
                "total_pedidos": len(pedidos_processados),
                "pedidos": pedidos_processados
            }, f, indent=2, ensure_ascii=False)
        
        print(f"  ✅ JSON: {json_file}")
        
        # 5. Exportar CSV
        csv_file = "pedidos_correios_exportacao.csv"
        if pedidos_processados:
            with open(csv_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=pedidos_processados[0].keys())
                writer.writeheader()
                writer.writerows(pedidos_processados)
            
            print(f"  ✅ CSV: {csv_file}")
        
        # 6. Tentar exportar Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            excel_file = "pedidos_correios_exportacao.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pedidos Correios"
            
            # Header
            headers = list(pedidos_processados[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Dados
            for row_idx, pedido in enumerate(pedidos_processados, 2):
                for col_idx, (key, value) in enumerate(pedido.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-ajustar colunas
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            print(f"  ✅ Excel: {excel_file}")
        
        except ImportError:
            print(f"  ⚠️  Excel: openpyxl não instalado (execute: pip install openpyxl)")
        
        # 7. Resumo
        print("\n" + "="*130)
        print("5. RESUMO")
        print("="*130 + "\n")
        
        print(f"  Total de pedidos: {len(pedidos_processados)}")
        print(f"  Período: {min(p['data_pedido'] for p in pedidos_processados)} a {max(p['data_pedido'] for p in pedidos_processados)}")
        
        valor_total = sum(float(p['valor_total'] or 0) for p in pedidos_processados)
        print(f"  Valor total: R$ {valor_total:,.2f}")
        
        clientes_unicos = len(set(p['cliente_nome'] for p in pedidos_processados))
        print(f"  Clientes únicos: {clientes_unicos}")
        
        print(f"\n  ✅ Exportação concluída!\n")
        
        return 0
        
    except Exception as e:
        print(f"  ❌ Erro: {e}\n")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
